"""Agent Output Capture: State diff -> DeepEval actual_output.

Captures kernel command output by diffing state before/after execution.
State diff feeds correctness metrics (ToolCorrectness, TaskCompletion).
Agent trace feeds faithfulness metrics (GEval protocol adherence).

Hybrid approach: both are captured, different metrics use different sources.
"""

import json
import os
import shutil
from pathlib import Path


def snapshot_state(repo_path, snapshot_dir):
    """Copy all state files (JSON, xlsx) from a repo to a snapshot directory.

    Args:
        repo_path: Root path of the repo to snapshot.
        snapshot_dir: Directory to copy state files into.

    Returns:
        list of relative paths that were copied.
    """
    repo = Path(repo_path)
    snap = Path(snapshot_dir)
    snap.mkdir(parents=True, exist_ok=True)

    copied = []
    patterns = ["**/*.json", "**/*.xlsx"]

    state_dirs = [
        repo / ".claude" / "state",
        repo / ".claude" / "protocols",
        repo / ".claude" / "lessons",
    ]

    for state_dir in state_dirs:
        if not state_dir.exists():
            continue
        for pattern in patterns:
            for src_file in state_dir.glob(pattern):
                rel = src_file.relative_to(repo)
                dst = snap / rel
                dst.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(str(src_file), str(dst))
                copied.append(str(rel))

    return sorted(copied)


def _load_json_safe(path):
    """Load a JSON file, returning None on failure."""
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, FileNotFoundError, OSError):
        return None


def _diff_json(before_obj, after_obj):
    """Compute a structured diff between two JSON objects.

    Returns dict of changed keys with before/after values.
    Only includes keys that actually changed.
    """
    if before_obj is None and after_obj is None:
        return {}
    if before_obj is None:
        return {"_added": after_obj}
    if after_obj is None:
        return {"_removed": before_obj}

    diff = {}
    all_keys = set(list(before_obj.keys()) + list(after_obj.keys())) if isinstance(before_obj, dict) and isinstance(after_obj, dict) else set()

    if not all_keys:
        if before_obj != after_obj:
            return {"_before": before_obj, "_after": after_obj}
        return {}

    for key in sorted(all_keys):
        before_val = before_obj.get(key)
        after_val = after_obj.get(key)
        if before_val != after_val:
            diff[key] = {"before": before_val, "after": after_val}

    return diff


def _diff_xlsx(before_path, after_path):
    """Compute cell-level diff between two xlsx files using openpyxl.

    Returns dict mapping "sheet:row_col" to before/after values.
    Returns empty dict if openpyxl is not available.
    """
    try:
        import openpyxl
    except ImportError:
        return {"_error": "openpyxl not installed"}

    if not os.path.exists(before_path) and not os.path.exists(after_path):
        return {}

    if not os.path.exists(before_path):
        wb_after = openpyxl.load_workbook(str(after_path), data_only=True)
        result = {}
        for sheet_name in wb_after.sheetnames:
            ws = wb_after[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        key = f"{sheet_name}:row_{cell.row}_col_{cell.column}"
                        result[key] = {"before": None, "after": cell.value}
        return result

    if not os.path.exists(after_path):
        return {"_removed": str(before_path)}

    wb_before = openpyxl.load_workbook(str(before_path), data_only=True)
    wb_after = openpyxl.load_workbook(str(after_path), data_only=True)

    diff = {}
    all_sheets = set(wb_before.sheetnames + wb_after.sheetnames)

    for sheet_name in sorted(all_sheets):
        ws_before = wb_before[sheet_name] if sheet_name in wb_before.sheetnames else None
        ws_after = wb_after[sheet_name] if sheet_name in wb_after.sheetnames else None

        if ws_before is None:
            for row in ws_after.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        key = f"{sheet_name}:row_{cell.row}_col_{cell.column}"
                        diff[key] = {"before": None, "after": cell.value}
            continue

        if ws_after is None:
            for row in ws_before.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        key = f"{sheet_name}:row_{cell.row}_col_{cell.column}"
                        diff[key] = {"before": cell.value, "after": None}
            continue

        max_row = max(ws_before.max_row or 0, ws_after.max_row or 0)
        max_col = max(ws_before.max_column or 0, ws_after.max_column or 0)

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                val_before = ws_before.cell(row=row, column=col).value
                val_after = ws_after.cell(row=row, column=col).value
                if val_before != val_after:
                    key = f"{sheet_name}:row_{row}_col_{col}"
                    diff[key] = {"before": val_before, "after": val_after}

    return diff


def diff_states(before_dir, after_dir):
    """Compute diff between two state snapshots.

    Args:
        before_dir: Path to the before-execution snapshot.
        after_dir: Path to the after-execution snapshot.

    Returns:
        dict with "state_diff" (JSON changes) and "xlsx_diff" (spreadsheet changes).
    """
    before_path = Path(before_dir)
    after_path = Path(after_dir)

    all_files = set()
    for d in [before_path, after_path]:
        if d.exists():
            for f in d.rglob("*"):
                if f.is_file():
                    all_files.add(str(f.relative_to(d)))

    state_diff = {}
    xlsx_diff = {}

    for rel_path in sorted(all_files):
        bf = before_path / rel_path
        af = after_path / rel_path

        if rel_path.endswith(".json"):
            before_obj = _load_json_safe(bf) if bf.exists() else None
            after_obj = _load_json_safe(af) if af.exists() else None
            file_diff = _diff_json(before_obj, after_obj)
            if file_diff:
                state_diff[rel_path] = file_diff

        elif rel_path.endswith(".xlsx"):
            file_diff = _diff_xlsx(str(bf), str(af))
            if file_diff:
                xlsx_diff[rel_path] = file_diff

    return {"state_diff": state_diff, "xlsx_diff": xlsx_diff}


def capture_actual_output(before_dir, after_dir):
    """Produce actual_output dict suitable for DeepEval's LLMTestCase.

    Args:
        before_dir: Path to the before-execution snapshot.
        after_dir: Path to the after-execution snapshot.

    Returns:
        dict with "files_changed", "state_diff", "xlsx_diff" keys.
    """
    diff = diff_states(before_dir, after_dir)

    files_changed = sorted(
        set(list(diff["state_diff"].keys()) + list(diff["xlsx_diff"].keys()))
    )

    return {
        "files_changed": files_changed,
        "state_diff": diff["state_diff"],
        "xlsx_diff": diff["xlsx_diff"],
    }
